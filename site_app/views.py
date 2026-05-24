import json
import csv
from datetime import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.conf import settings

from .models import Alignment, FeatureCapture, PassingPlace, FeaturePhoto, PassingPlacePhoto
from .utils import (
    load_alignment_from_dxf,
    get_available_dxf_files,
    gps_to_projected,
    chainage_to_gps,
    get_alignment_gps_line,
    get_next_pp_id,
)


# -----------------------------
# Helper — load alignment or 404
# -----------------------------
def get_alignment_data(alignment):
    """Load DXF data for an alignment. Returns None if file missing."""
    return load_alignment_from_dxf(alignment.dxf_file)


# -----------------------------
# Auth views
# -----------------------------
def login_view(request):
    if request.user.is_authenticated:
        return redirect("dashboard")

    error = None
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("dashboard")
        else:
            error = "Invalid username or password."

    return render(request, "login.html", {"error": error})


def logout_view(request):
    logout(request)
    return redirect("login")


# -----------------------------
# Dashboard — site selector
# -----------------------------
@login_required
def dashboard(request):
    alignments = Alignment.objects.filter(active=True)
    return render(request, "dashboard.html", {"alignments": alignments})


# -----------------------------
# Phone GPS
# -----------------------------
@login_required
def phone_gps(request, alignment_id):
    alignment = get_object_or_404(Alignment, id=alignment_id, active=True)
    data      = get_alignment_data(alignment)
    if data is None:
        messages.error(request, f"Could not load DXF file: {alignment.dxf_file}")
        return redirect("dashboard")

    gps_line = get_alignment_gps_line(data["points"])

    return render(request, "phone_gps.html", {
        "alignment":  alignment,
        "gps_line":   json.dumps(gps_line),
        "total":      round(data["total_length"], 3),
    })


# -----------------------------
# Capture feature
# -----------------------------
@login_required
def capture(request, alignment_id):
    alignment = get_object_or_404(Alignment, id=alignment_id, active=True)
    data      = get_alignment_data(alignment)
    if data is None:
        messages.error(request, f"Could not load DXF file: {alignment.dxf_file}")
        return redirect("dashboard")

    gps_line      = get_alignment_gps_line(data["points"])
    feature_types = FeatureCapture.FEATURE_TYPES

    if request.method == "POST":
        entry_method = request.POST.get("entry_method", "GPS")

        if entry_method == "Manual":
            try:
                chainage  = float(request.POST.get("manual_chainage", 0))
                projected = chainage_to_gps(data["points"], data["cum_dist"], chainage)
                projected["distance_from_alignment"] = 0.0
            except (TypeError, ValueError):
                messages.error(request, "Invalid chainage value.")
                return redirect("capture", alignment_id=alignment_id)
        else:
            try:
                lat = float(request.POST.get("latitude"))
                lon = float(request.POST.get("longitude"))
            except (TypeError, ValueError):
                messages.error(request, "No GPS location captured — please get your location first.")
                return render(request, "capture.html", {
                    "alignment":     alignment,
                    "gps_line":      json.dumps(gps_line),
                    "feature_types": feature_types,
                    "total":         round(data["total_length"], 3),
                })
            projected = gps_to_projected(data["points"], data["cum_dist"], lat, lon)

        feature_type = request.POST.get("feature_type", "")
        custom_type  = request.POST.get("custom_feature_type", "")
        photos       = request.FILES.getlist("photos")

        feature = FeatureCapture.objects.create(
            alignment             = alignment,
            feature_type          = feature_type,
            custom_feature_type   = custom_type,
            side                  = request.POST.get("side", "NA"),
            condition             = request.POST.get("condition", "GOOD"),
            offset_from_edge_m    = float(request.POST.get("offset_from_edge_m", 0) or 0),
            distance_along_edge_m = float(request.POST.get("distance_along_edge_m", 0) or 0),
            notes                 = request.POST.get("notes", ""),
            chainage_m            = projected["chainage"],
            distance_from_alignment_m = projected["distance_from_alignment"],
            latitude              = projected["latitude"],
            longitude             = projected["longitude"],
            easting               = projected["easting"],
            northing              = projected["northing"],
            gps_accuracy_m        = float(request.POST.get("gps_accuracy", 0) or 0) if entry_method == "GPS" else None,
            entry_method          = entry_method,
            captured_by           = request.user,
        )

        for photo in photos:
            FeaturePhoto.objects.create(feature=feature, photo=photo)

        messages.success(
            request,
            f"✅ Saved F{feature.id:03d} [{entry_method}]: {custom_type or feature_type} at chainage {projected['chainage']}m"
        )
        return redirect("capture", alignment_id=alignment_id)

    return render(request, "capture.html", {
        "alignment":         alignment,
        "gps_line":          json.dumps(gps_line),
        "feature_types":     feature_types,
        "total":             round(data["total_length"], 3),
    })


# -----------------------------
# Passing places
# -----------------------------
@login_required
def passing_places(request, alignment_id):
    alignment = get_object_or_404(Alignment, id=alignment_id, active=True)
    data      = get_alignment_data(alignment)
    if data is None:
        messages.error(request, f"Could not load DXF file: {alignment.dxf_file}")
        return redirect("dashboard")

    gps_line  = get_alignment_gps_line(data["points"])
    next_id   = get_next_pp_id(alignment)
    pp_list   = PassingPlace.objects.filter(alignment=alignment)

    if request.method == "POST":
        entry_method = request.POST.get("entry_method", "GPS")

        if entry_method == "Manual":
            try:
                chainage  = float(request.POST.get("manual_chainage", 0))
                projected = chainage_to_gps(data["points"], data["cum_dist"], chainage)
                projected["distance_from_alignment"] = 0.0
            except (TypeError, ValueError):
                messages.error(request, "Invalid chainage value.")
                return redirect("passing_places", alignment_id=alignment_id)
        else:
            try:
                lat = float(request.POST.get("latitude"))
                lon = float(request.POST.get("longitude"))
            except (TypeError, ValueError):
                messages.error(request, "No GPS location — please get your location first.")
                return render(request, "passing_places.html", {
                    "alignment": alignment,
                    "gps_line":  json.dumps(gps_line),
                    "next_id":   next_id,
                    "pp_list":   pp_list,
                    "total":     round(data["total_length"], 3),
                })
            projected = gps_to_projected(data["points"], data["cum_dist"], lat, lon)

        photos = request.FILES.getlist("photos")

        pp = PassingPlace.objects.create(
            alignment      = alignment,
            pp_id          = next_id,
            side           = request.POST.get("side", "LHS"),
            status         = request.POST.get("status", "Existing"),
            mid_chainage_m = projected["chainage"],
            mid_latitude   = projected["latitude"],
            mid_longitude  = projected["longitude"],
            mid_easting    = projected["easting"],
            mid_northing   = projected["northing"],
            width_m        = float(request.POST.get("width_m", 0) or 0),
            length_m       = float(request.POST.get("length_m", 0) or 0),
            notes          = request.POST.get("notes", ""),
            gps_accuracy_m = float(request.POST.get("gps_accuracy", 0) or 0) if entry_method == "GPS" else None,
            entry_method   = entry_method,
            captured_by    = request.user,
        )

        for photo in photos:
            PassingPlacePhoto.objects.create(passing_place=pp, photo=photo)

        messages.success(
            request,
            f"✅ Saved {next_id} [{entry_method}]: chainage {projected['chainage']}m"
        )
        return redirect("passing_places", alignment_id=alignment_id)

    return render(request, "passing_places.html", {
        "alignment": alignment,
        "gps_line":  json.dumps(gps_line),
        "next_id":   next_id,
        "pp_list":   pp_list,
        "total":     round(data["total_length"], 3),
    })


# -----------------------------
# View captured points
# -----------------------------
@login_required
def view_points(request, alignment_id):
    alignment = get_object_or_404(Alignment, id=alignment_id, active=True)
    data      = get_alignment_data(alignment)
    if data is None:
        messages.error(request, f"Could not load DXF file: {alignment.dxf_file}")
        return redirect("dashboard")

    gps_line = get_alignment_gps_line(data["points"])
    features = FeatureCapture.objects.filter(alignment=alignment)
    pp_list  = PassingPlace.objects.filter(alignment=alignment)

    return render(request, "view_points.html", {
        "alignment": alignment,
        "gps_line":  json.dumps(gps_line),
        "features":  features,
        "pp_list":   pp_list,
    })


# -----------------------------
# Tools — Chainage ↔ GPS
# -----------------------------
@login_required
def tools(request, alignment_id):
    alignment = get_object_or_404(Alignment, id=alignment_id, active=True)
    data      = get_alignment_data(alignment)
    if data is None:
        messages.error(request, f"Could not load DXF file: {alignment.dxf_file}")
        return redirect("dashboard")

    gps_line        = get_alignment_gps_line(data["points"])
    ch_result       = None
    gps_result      = None

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "chainage_to_gps":
            try:
                chainage  = float(request.POST.get("chainage", 0))
                ch_result = chainage_to_gps(data["points"], data["cum_dist"], chainage)
            except (ValueError, TypeError):
                messages.error(request, "Invalid chainage value.")

        elif action == "gps_to_chainage":
            try:
                lat        = float(request.POST.get("latitude", 0))
                lon        = float(request.POST.get("longitude", 0))
                gps_result = gps_to_projected(data["points"], data["cum_dist"], lat, lon)
            except (ValueError, TypeError):
                messages.error(request, "Invalid coordinates.")

    return render(request, "tools.html", {
        "alignment":  alignment,
        "gps_line":   json.dumps(gps_line),
        "total":      round(data["total_length"], 3),
        "ch_result":  ch_result,
        "gps_result": gps_result,
    })


# -----------------------------
# AJAX — GPS to chainage
# -----------------------------
@login_required
def api_gps_to_chainage(request, alignment_id):
    alignment = get_object_or_404(Alignment, id=alignment_id, active=True)
    data      = get_alignment_data(alignment)
    if data is None:
        return JsonResponse({"error": "DXF file not found"}, status=404)

    try:
        body   = json.loads(request.body)
        lat    = float(body["latitude"])
        lon    = float(body["longitude"])
        result = gps_to_projected(data["points"], data["cum_dist"], lat, lon)
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# -----------------------------
# AJAX — nearest features to chainage
# -----------------------------
@login_required
def api_nearest_features(request, alignment_id):
    alignment = get_object_or_404(Alignment, id=alignment_id, active=True)

    try:
        body     = json.loads(request.body)
        chainage = float(body["chainage"])
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)

    # Features
    features_before = FeatureCapture.objects.filter(
        alignment=alignment,
        chainage_m__lte=chainage
    ).order_by("-chainage_m").first()

    features_ahead = FeatureCapture.objects.filter(
        alignment=alignment,
        chainage_m__gt=chainage
    ).order_by("chainage_m").first()

    # Passing places
    pp_before = PassingPlace.objects.filter(
        alignment=alignment,
        mid_chainage_m__lte=chainage
    ).order_by("-mid_chainage_m").first()

    pp_ahead = PassingPlace.objects.filter(
        alignment=alignment,
        mid_chainage_m__gt=chainage
    ).order_by("mid_chainage_m").first()

    def feature_dict(f, current_chainage):
        if f is None:
            return None
        dist = abs(f.chainage_m - current_chainage)
        return {
            "chainage":     round(f.chainage_m, 3),
            "distance":     round(dist, 1),
            "label":        f.get_feature_label(),
            "condition":    f.condition,
            "side":         f.side,
        }

    def pp_dict(pp, current_chainage):
        if pp is None:
            return None
        dist = abs(pp.mid_chainage_m - current_chainage)
        return {
            "chainage":  round(pp.mid_chainage_m, 3),
            "distance":  round(dist, 1),
            "pp_id":     pp.pp_id,
            "side":      pp.side,
            "status":    pp.status,
        }

    return JsonResponse({
        "feature_before":  feature_dict(features_before, chainage),
        "feature_ahead":   feature_dict(features_ahead,  chainage),
        "pp_before":       pp_dict(pp_before, chainage),
        "pp_ahead":        pp_dict(pp_ahead,  chainage),
    })
@login_required
def api_chainage_to_gps(request, alignment_id):
    alignment = get_object_or_404(Alignment, id=alignment_id, active=True)
    data      = get_alignment_data(alignment)
    if data is None:
        return JsonResponse({"error": "DXF file not found"}, status=404)

    try:
        body     = json.loads(request.body)
        chainage = float(body["chainage"])
        result   = chainage_to_gps(data["points"], data["cum_dist"], chainage)
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# -----------------------------
# AJAX — save feature capture
# -----------------------------
@login_required
def api_capture(request, alignment_id):
    alignment = get_object_or_404(Alignment, id=alignment_id, active=True)
    data      = get_alignment_data(alignment)
    if data is None:
        return JsonResponse({"error": "DXF file not found"}, status=404)

    try:
        body        = json.loads(request.body)
        lat         = float(body["latitude"])
        lon         = float(body["longitude"])
        projected   = gps_to_projected(data["points"], data["cum_dist"], lat, lon)

        feature_type = body.get("feature_type", "")
        custom       = body.get("custom_feature_type", "")

        FeatureCapture.objects.create(
            alignment             = alignment,
            feature_type          = feature_type,
            custom_feature_type   = custom,
            side                  = body.get("side", "NA"),
            condition             = body.get("condition", "GOOD"),
            offset_from_edge_m    = float(body.get("offset_from_edge_m", 0)),
            distance_along_edge_m = float(body.get("distance_along_edge_m", 0)),
            notes                 = body.get("notes", ""),
            chainage_m            = projected["chainage"],
            distance_from_alignment_m = projected["distance_from_alignment"],
            latitude              = projected["latitude"],
            longitude             = projected["longitude"],
            easting               = projected["easting"],
            northing              = projected["northing"],
            captured_by           = request.user,
        )
        return JsonResponse({"success": True, "chainage": projected["chainage"]})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# -----------------------------
# AJAX — save passing place
# -----------------------------
@login_required
def api_passing_place(request, alignment_id):
    alignment = get_object_or_404(Alignment, id=alignment_id, active=True)
    data      = get_alignment_data(alignment)
    if data is None:
        return JsonResponse({"error": "DXF file not found"}, status=404)

    try:
        body      = json.loads(request.body)
        lat       = float(body["latitude"])
        lon       = float(body["longitude"])
        projected = gps_to_projected(data["points"], data["cum_dist"], lat, lon)
        next_id   = get_next_pp_id(alignment)

        PassingPlace.objects.create(
            alignment      = alignment,
            pp_id          = next_id,
            side           = body.get("side", "LHS"),
            status         = body.get("status", "Existing"),
            mid_chainage_m = projected["chainage"],
            mid_latitude   = projected["latitude"],
            mid_longitude  = projected["longitude"],
            mid_easting    = projected["easting"],
            mid_northing   = projected["northing"],
            width_m        = float(body.get("width_m", 0)),
            length_m       = float(body.get("length_m", 0)),
            notes          = body.get("notes", ""),
            gps_accuracy_m = float(body.get("gps_accuracy_m", 0) or 0),
            captured_by    = request.user,
        )
        return JsonResponse({"success": True, "pp_id": next_id, "chainage": projected["chainage"]})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


# -----------------------------
# AJAX — add photos to existing feature
# -----------------------------
@login_required
def api_add_feature_photos(request, feature_id):
    feature = get_object_or_404(FeatureCapture, id=feature_id)
    photos  = request.FILES.getlist("photos")
    if not photos:
        return JsonResponse({"error": "No photos received"}, status=400)
    saved = []
    for photo in photos:
        fp = FeaturePhoto.objects.create(feature=feature, photo=photo)
        saved.append(fp.photo.url)
    return JsonResponse({"success": True, "urls": saved, "count": len(saved)})


# -----------------------------
# AJAX — add photos to existing passing place
# -----------------------------
@login_required
def api_add_pp_photos(request, pp_id):
    pp     = get_object_or_404(PassingPlace, id=pp_id)
    photos = request.FILES.getlist("photos")
    if not photos:
        return JsonResponse({"error": "No photos received"}, status=400)
    saved = []
    for photo in photos:
        pp_photo = PassingPlacePhoto.objects.create(passing_place=pp, photo=photo)
        saved.append(pp_photo.photo.url)
    return JsonResponse({"success": True, "urls": saved, "count": len(saved)})
@login_required
def admin_export(request):
    if not request.user.is_staff:
        return redirect("dashboard")

    from pathlib import Path
    MAX_VOLUME_BYTES = 100 * 1024 * 1024  # 100MB per volume

    alignments = Alignment.objects.filter(active=True)
    alignment_data = []

    for a in alignments:
        feature_photos = sum(f.photos.count() for f in a.features.all())
        pp_photos      = sum(pp.photos.count() for pp in a.passing_places.all())

        # Calculate photo volumes — features
        feature_photo_list = []
        for f in a.features.all():
            for fp in f.photos.all():
                try:
                    photo_path = Path(settings.MEDIA_ROOT) / fp.photo.name
                    if photo_path.exists():
                        feature_photo_list.append(photo_path)
                except Exception:
                    pass

        # Calculate photo volumes — passing places
        pp_photo_list = []
        for pp in a.passing_places.all():
            for pp_photo in pp.photos.all():
                try:
                    photo_path = Path(settings.MEDIA_ROOT) / pp_photo.photo.name
                    if photo_path.exists():
                        pp_photo_list.append(photo_path)
                except Exception:
                    pass

        def make_volumes(photo_list):
            volumes = []
            current_vol  = []
            current_size = 0
            for p in photo_list:
                size = p.stat().st_size
                if current_size + size > MAX_VOLUME_BYTES and current_vol:
                    volumes.append({
                        "number":  len(volumes) + 1,
                        "count":   len(current_vol),
                        "size_mb": round(current_size / 1024 / 1024, 1)
                    })
                    current_vol  = []
                    current_size = 0
                current_vol.append(p)
                current_size += size
            if current_vol:
                volumes.append({
                    "number":  len(volumes) + 1,
                    "count":   len(current_vol),
                    "size_mb": round(current_size / 1024 / 1024, 1)
                })
            return volumes

        alignment_data.append({
            "alignment":       a,
            "feature_count":   a.features.count(),
            "pp_count":        a.passing_places.count(),
            "photo_count":     feature_photos + pp_photos,
            "feature_volumes": make_volumes(feature_photo_list),
            "pp_volumes":      make_volumes(pp_photo_list),
        })

    return render(request, "admin_export.html", {"alignment_data": alignment_data})


@login_required
def export_excel_only(request, alignment_id):
    if not request.user.is_staff:
        return redirect("dashboard")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment as XLAlign
    from io import BytesIO
    from pathlib import Path

    alignment = get_object_or_404(Alignment, id=alignment_id)
    features  = FeatureCapture.objects.filter(alignment=alignment)
    pp_list   = PassingPlace.objects.filter(alignment=alignment)
    stem      = alignment.dxf_file.replace(".dxf", "")

    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Features"
    header_fill = PatternFill("solid", fgColor="0d6efd")
    header_font = Font(bold=True, color="FFFFFF")

    feature_headers = [
        "ID", "Entry Method", "Feature Type", "Side", "Condition",
        "Offset from Edge (m)", "Distance Along Edge (m)",
        "Chainage (m)", "Distance from Alignment (m)",
        "Latitude", "Longitude", "Easting", "Northing",
        "GPS Accuracy (m)", "Photo", "Notes", "Captured By", "Captured At"
    ]
    for col, header in enumerate(feature_headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = XLAlign(horizontal="center")

    for f in features:
        photo_names = ", ".join([Path(fp.photo.name).name for fp in f.photos.all()])
        ws1.append([
            f"F{f.id:03d}", f.entry_method, f.get_feature_label(), f.side, f.condition,
            f.offset_from_edge_m, f.distance_along_edge_m,
            f.chainage_m, f.distance_from_alignment_m,
            f.latitude, f.longitude, f.easting, f.northing,
            f.gps_accuracy_m, photo_names, f.notes,
            f.captured_by.username if f.captured_by else "",
            f.captured_at.strftime("%Y-%m-%d %H:%M:%S")
        ])

    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws2 = wb.create_sheet("Passing Places")
    pp_headers = [
        "ID", "Entry Method", "PP ID", "Side", "Status",
        "Mid Chainage (m)", "Mid Latitude", "Mid Longitude",
        "Mid Easting", "Mid Northing", "Width (m)", "Length (m)",
        "GPS Accuracy (m)", "Photo", "Notes", "Captured By", "Captured At"
    ]
    for col, header in enumerate(pp_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = XLAlign(horizontal="center")

    for pp in pp_list:
        photo_names = ", ".join([Path(pp_photo.photo.name).name for pp_photo in pp.photos.all()])
        ws2.append([
            pp.id, pp.entry_method, pp.pp_id, pp.side, pp.status,
            pp.mid_chainage_m, pp.mid_latitude, pp.mid_longitude,
            pp.mid_easting, pp.mid_northing,
            pp.width_m, pp.length_m,
            pp.gps_accuracy_m, photo_names, pp.notes,
            pp.captured_by.username if pp.captured_by else "",
            pp.captured_at.strftime("%Y-%m-%d %H:%M:%S")
        ])

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws3 = wb.create_sheet("Summary")
    ws3.append(["Alignment",            stem])
    ws3.append(["Total Features",       features.count()])
    ws3.append(["Total Passing Places", pp_list.count()])
    ws3.append(["Export Date",          datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    filename = f"{stem}_data_{datetime.now().strftime('%Y%m%d')}.xlsx"
    buffer   = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_photos_volume(request, alignment_id, photo_type, volume):
    if not request.user.is_staff:
        return redirect("dashboard")

    import zipfile
    from io import BytesIO
    from pathlib import Path

    MAX_VOLUME_BYTES = 100 * 1024 * 1024  # 100MB

    alignment = get_object_or_404(Alignment, id=alignment_id)
    stem      = alignment.dxf_file.replace(".dxf", "")

    # Collect photos based on type
    all_photos = []
    if photo_type == "features":
        for f in alignment.features.all():
            for fp in f.photos.all():
                try:
                    photo_path = Path(settings.MEDIA_ROOT) / fp.photo.name
                    if photo_path.exists():
                        all_photos.append((photo_path, Path(fp.photo.name).name))
                except Exception:
                    pass
    else:
        for pp in alignment.passing_places.all():
            for pp_photo in pp.photos.all():
                try:
                    photo_path = Path(settings.MEDIA_ROOT) / pp_photo.photo.name
                    if photo_path.exists():
                        all_photos.append((photo_path, Path(pp_photo.photo.name).name))
                except Exception:
                    pass

    # Split into volumes
    volumes      = []
    current_vol  = []
    current_size = 0
    for item in all_photos:
        size = item[0].stat().st_size
        if current_size + size > MAX_VOLUME_BYTES and current_vol:
            volumes.append(current_vol)
            current_vol  = []
            current_size = 0
        current_vol.append(item)
        current_size += size
    if current_vol:
        volumes.append(current_vol)

    vol_index = volume - 1
    if vol_index < 0 or vol_index >= len(volumes):
        messages.error(request, "Volume not found.")
        return redirect("admin_export")

    selected = volumes[vol_index]
    type_label = "features" if photo_type == "features" else "passing_places"
    filename   = f"{stem}_{type_label}_photos_vol{volume}_{datetime.now().strftime('%Y%m%d')}.zip"

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for photo_path, photo_name in selected:
            zf.write(photo_path, f"photos/{type_label}/{photo_name}")

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
    if not request.user.is_staff:
        return redirect("dashboard")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment as XLAlign
    from io import BytesIO
    import zipfile
    from pathlib import Path

    alignment = get_object_or_404(Alignment, id=alignment_id)
    features  = FeatureCapture.objects.filter(alignment=alignment)
    pp_list   = PassingPlace.objects.filter(alignment=alignment)
    stem      = alignment.dxf_file.replace(".dxf", "")
    datestamp = datetime.now().strftime("%Y%m%d")
    folder    = f"{stem}_export_{datestamp}"

    # -----------------------------
    # Build Excel
    # -----------------------------
    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Features"

    header_fill = PatternFill("solid", fgColor="0d6efd")
    header_font = Font(bold=True, color="FFFFFF")

    feature_headers = [
        "ID", "Entry Method", "Feature Type", "Side", "Condition",
        "Offset from Edge (m)", "Distance Along Edge (m)",
        "Chainage (m)", "Distance from Alignment (m)",
        "Latitude", "Longitude", "Easting", "Northing",
        "GPS Accuracy (m)", "Photo", "Notes", "Captured By", "Captured At"
    ]
    for col, header in enumerate(feature_headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = XLAlign(horizontal="center")

    for f in features:
        photo_names = ", ".join([Path(fp.photo.name).name for fp in f.photos.all()])
        ws1.append([
            f"F{f.id:03d}", f.entry_method, f.get_feature_label(), f.side, f.condition,
            f.offset_from_edge_m, f.distance_along_edge_m,
            f.chainage_m, f.distance_from_alignment_m,
            f.latitude, f.longitude, f.easting, f.northing,
            f.gps_accuracy_m, photo_names,
            f.notes,
            f.captured_by.username if f.captured_by else "",
            f.captured_at.strftime("%Y-%m-%d %H:%M:%S")
        ])

    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws2 = wb.create_sheet("Passing Places")
    pp_headers = [
        "ID", "PP ID", "Side", "Status",
        "Mid Chainage (m)", "Mid Latitude", "Mid Longitude",
        "Mid Easting", "Mid Northing",
        "Width (m)", "Length (m)",
        "GPS Accuracy (m)", "Photo", "Notes", "Captured By", "Captured At"
    ]
    for col, header in enumerate(pp_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = XLAlign(horizontal="center")

    for pp in pp_list:
        photo_names = ", ".join([Path(pp_photo.photo.name).name for pp_photo in pp.photos.all()])
        ws2.append([
            pp.id, pp.pp_id, pp.side, pp.status,
            pp.mid_chainage_m, pp.mid_latitude, pp.mid_longitude,
            pp.mid_easting, pp.mid_northing,
            pp.width_m, pp.length_m,
            pp.gps_accuracy_m, photo_names,
            pp.notes,
            pp.captured_by.username if pp.captured_by else "",
            pp.captured_at.strftime("%Y-%m-%d %H:%M:%S")
        ])

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws3 = wb.create_sheet("Summary")
    ws3.append(["Alignment",            stem])
    ws3.append(["Total Features",       features.count()])
    ws3.append(["Total Passing Places", pp_list.count()])
    ws3.append(["Export Date",          datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # -----------------------------
    # Build ZIP
    # -----------------------------
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:

        # Add Excel
        zf.writestr(
            f"{folder}/{stem}_export_{datestamp}.xlsx",
            excel_buffer.getvalue()
        )

        # Add feature photos
        for f in features:
            for fp in f.photos.all():
                photo_path = Path(settings.MEDIA_ROOT) / fp.photo.name
                if photo_path.exists():
                    zf.write(
                        photo_path,
                        f"{folder}/photos/features/{Path(fp.photo.name).name}"
                    )

        # Add passing place photos
        for pp in pp_list:
            for pp_photo in pp.photos.all():
                photo_path = Path(settings.MEDIA_ROOT) / pp_photo.photo.name
                if photo_path.exists():
                    zf.write(
                        photo_path,
                        f"{folder}/photos/passing_places/{Path(pp_photo.photo.name).name}"
                    )

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{folder}.zip"'
    return response


# -----------------------------
# Edit feature
# -----------------------------
@login_required
def edit_feature(request, feature_id):
    feature   = get_object_or_404(FeatureCapture, id=feature_id)
    alignment = feature.alignment

    if request.method == "POST":
        feature.feature_type          = request.POST.get("feature_type", feature.feature_type)
        feature.custom_feature_type   = request.POST.get("custom_feature_type", "")
        feature.side                  = request.POST.get("side", feature.side)
        feature.condition             = request.POST.get("condition", feature.condition)
        feature.offset_from_edge_m    = float(request.POST.get("offset_from_edge_m", 0) or 0)
        feature.distance_along_edge_m = float(request.POST.get("distance_along_edge_m", 0) or 0)
        feature.notes                 = request.POST.get("notes", "")
        feature.save()
        messages.success(request, f"✅ F{feature.id:03d} updated successfully.")
        return redirect("view_points", alignment_id=alignment.id)

    return render(request, "edit_feature.html", {
        "feature":      feature,
        "alignment":    alignment,
        "feature_types": FeatureCapture.FEATURE_TYPES,
    })


# -----------------------------
# Edit passing place
# -----------------------------
@login_required
def edit_passing_place(request, pp_id):
    pp        = get_object_or_404(PassingPlace, id=pp_id)
    alignment = pp.alignment

    if request.method == "POST":
        pp.side     = request.POST.get("side", pp.side)
        pp.status   = request.POST.get("status", pp.status)
        pp.width_m  = float(request.POST.get("width_m", 0) or 0)
        pp.length_m = float(request.POST.get("length_m", 0) or 0)
        pp.notes    = request.POST.get("notes", "")
        pp.save()
        messages.success(request, f"✅ {pp.pp_id} updated successfully.")
        return redirect("view_points", alignment_id=alignment.id)

    return render(request, "edit_passing_place.html", {
        "pp":        pp,
        "alignment": alignment,
    })


# -----------------------------
# Delete passing place
# -----------------------------
@login_required
def delete_passing_place(request, pp_id):
    pp = get_object_or_404(PassingPlace, id=pp_id)
    alignment_id = pp.alignment.id
    pp.delete()
    messages.success(request, "Passing place deleted.")
    return redirect("view_points", alignment_id=alignment_id)


# -----------------------------
# Delete feature
# -----------------------------
@login_required
def delete_feature(request, feature_id):
    feature = get_object_or_404(FeatureCapture, id=feature_id)
    alignment_id = feature.alignment.id
    feature.delete()
    messages.success(request, "Feature deleted.")
    return redirect("view_points", alignment_id=alignment_id)
@login_required
def export_features_csv(request, alignment_id):
    alignment = get_object_or_404(Alignment, id=alignment_id, active=True)
    features  = FeatureCapture.objects.filter(alignment=alignment)
    stem      = alignment.dxf_file.replace(".dxf", "")
    filename  = f"{stem}_points_{datetime.now().strftime('%Y%m%d')}.csv"

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        "ID", "Feature Type", "Side", "Condition",
        "Offset from Edge (m)", "Distance Along Edge (m)",
        "Chainage (m)", "Distance from Alignment (m)",
        "Latitude", "Longitude", "Easting", "Northing",
        "GPS Accuracy (m)", "Notes", "Captured By", "Captured At"
    ])
    for f in features:
        writer.writerow([
            f.id, f.get_feature_label(), f.side, f.condition,
            f.offset_from_edge_m, f.distance_along_edge_m,
            f.chainage_m, f.distance_from_alignment_m,
            f.latitude, f.longitude, f.easting, f.northing,
            f.gps_accuracy_m,
            f.notes, f.captured_by.username if f.captured_by else "",
            f.captured_at.strftime("%Y-%m-%d %H:%M:%S")
        ])

    return response


# -----------------------------
# Export passing places CSV
# -----------------------------
@login_required
def export_passing_places_csv(request, alignment_id):
    alignment = get_object_or_404(Alignment, id=alignment_id, active=True)
    pp_list   = PassingPlace.objects.filter(alignment=alignment)
    stem      = alignment.dxf_file.replace(".dxf", "")
    filename  = f"{stem}_passing_places_{datetime.now().strftime('%Y%m%d')}.csv"

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        "ID", "PP ID", "Side", "Status",
        "Mid Chainage (m)", "Mid Latitude", "Mid Longitude",
        "Mid Easting", "Mid Northing",
        "Width (m)", "Length (m)",
        "GPS Accuracy (m)", "Notes", "Captured By", "Captured At"
    ])
    for pp in pp_list:
        writer.writerow([
            pp.id, pp.pp_id, pp.side, pp.status,
            pp.mid_chainage_m, pp.mid_latitude, pp.mid_longitude,
            pp.mid_easting, pp.mid_northing,
            pp.width_m, pp.length_m,
            pp.gps_accuracy_m,
            pp.notes, pp.captured_by.username if pp.captured_by else "",
            pp.captured_at.strftime("%Y-%m-%d %H:%M:%S")
        ])

    return response
